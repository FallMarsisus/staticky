import openpyxl
from datetime import datetime, timedelta, timezone
import discord
from discord import app_commands, Embed
from PIL import Image, ImageDraw, ImageFont

def create_edt(heure:int, semestre:int, day:str) -> list[(int, str, int)]:

    days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    days_english = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    try:
        day_number_sem = days.index(day)
    except ValueError:
        day_number_sem = days_english.index(day)
    workbook = openpyxl.load_workbook('edt salles.xlsx')

    sheet = workbook['EDT']

    # Retrieve all coordinates where the cell value is 10
    coordinates = []
    
    # initialize the coords of the cell
    row = 3 + (heure - 8) + 11*day_number_sem
    column = 0
    for cell in sheet[row][1:]:
        column+=1
        val = cell.value
        if isEmptyRoom(val):
            try:
                roomName = sheet.cell(row=1, column=cell.column).value
                
                if roomName is not None:
                    freeTime = getFreeDuration(sheet, row, column, day_number_sem, ("R"==roomName[0]))
                    coordinates.append((roomName, sheet.cell(row=2, column=cell.column).value, freeTime))
                    
            except Exception as e:
                print(f"\033[91mError: {e}\033[0m") # affiche en rouge l'erreur
                
    return coordinates


def isEmptyRoom(val:str) -> bool:
    return (val is None or ("0" not in val and "O" not in val and "S1" not in val))
    
def isThisDay(row:int, dayNumber: int) -> bool : 
	return 3 + 11*dayNumber <= row <= 13 + 11*dayNumber


def getFreeDuration(sheet, row:int, column:int, dayNumber:int, isScienceBat:bool) -> int:
    i = 0
    val = sheet[row][column].value
    while isEmptyRoom(val) and isThisDay(row, dayNumber):
        i += 1
        row += 1
        val = sheet[row][column].value
        
        
    #complète jusqu'à 21 si la salle est libre après
    if not isThisDay(row, dayNumber) and not isScienceBat:
        return i + 3
    
    return i


def set_value(type, hour) -> str:
    if type == "Amphi":
        return "🧪 Amphi"
    if type == "TP":
        return "🔬 TP"
    if type == "Info":
        return "💻 Info"
    if type == "Colle":
        return "📖 Colle"
    if type == "Classe":
        return "🎓 Classe"
    if type == "?":
        return "❓ Inconnu"
    else:
        return ""
    
def create_image(coordinates, hour, day):
    # Create a larger blank image with a custom background color
    img = Image.new('RGB', (1200, 5000), color=(60, 63, 65))
    draw = ImageDraw.Draw(img)

    # Load a less formal and slightly larger font
    font = ImageFont.truetype("font.ttf", 45)
    title_font = ImageFont.truetype("font.ttf", 55)
    
    # Title
    title = f"EDT : {hour}h - {day}"
    draw.text((600 - (draw.textbbox((0, 0), title, font=title_font)[2] - draw.textbbox((0, 0), title, font=title_font)[0]) // 2, 20), title, fill="white", font=title_font)

    # Description
    description = f"Liste des Salles disponibles à {hour}h"
    draw.text((600 - (draw.textbbox((0, 0), description, font=font)[2] - draw.textbbox((0, 0), description, font=font)[0]) // 2, 80), description, fill="white", font=font)

    # Draw the table header with a different color
    header = ["Salle", "Type", "Temps libre"]
    x_text = [200, 600, 1000]
    for i, h in enumerate(header):
        draw.text((x_text[i] - (draw.textbbox((0, 0), h, font=font)[2] - draw.textbbox((0, 0), h, font=font)[0]) // 2, 140), h, fill="white", font=font)

    # Define background colors for each type of room
    type_colors = {
        "Amphi": (255, 153, 153),
        "TP": (153, 204, 255),
        "Info": (255, 204, 204),
        "Colle": (255, 255, 153),
        "Classe": (255, 229, 153),
        "?": (224, 224, 224)
    }

    # Sort coordinates by room type
    coordinates.sort(key=lambda x: x[1])

    # Draw the coordinates in a table format with alternating row colors
    y_text = 200
    line_height = 45
    padding = 8
    for i, (room, room_type, freeTime) in enumerate(coordinates):
        if y_text + line_height + padding > 3000:
            break
        fill_color = type_colors.get(room_type, (150, 150, 150))
        
        # Draw rounded rectangle
        draw.rounded_rectangle([(0, y_text), (1200, y_text + line_height)], radius=20, fill=fill_color)
        
        text_y = y_text + (line_height - font.getbbox(room)[3]) // 2
        
        # Draw the room name
        draw.text((200 - (draw.textbbox((0, 0), room, font=font)[2] - draw.textbbox((0, 0), room, font=font)[0]) // 2, text_y), room, fill="black", font=font)
        
        # Draw the room type in darker gray
        draw.text((600 - (draw.textbbox((0, 0), set_value(room_type, hour), font=font)[2] - draw.textbbox((0, 0), set_value(room_type, hour), font=font)[0]) // 2, text_y), set_value(room_type, hour), fill="gray", font=font)
        
        # Draw the free time duration
        draw.text((1000 - (draw.textbbox((0, 0), f"{freeTime}h", font=font)[2] - draw.textbbox((0, 0), f"{freeTime}h", font=font)[0]) // 2, text_y), f"{freeTime}h", fill="black", font=font)
        
        y_text += line_height + padding

    # Resize the image to fit the content
    img = img.crop((0, 0, 1200, y_text))

    img.save("edt.png")
    
def parse_edt(coordinates, hour, day):
    embeds = []

    embed1 = discord.Embed(title=f"EDT : {hour}h - {day}",
                           url="https://staticky.marsisus.me",
                           description=f"Liste des Salles disponibles à {hour}h",
                           timestamp=datetime.now())

    for i in range(min(25, len(coordinates))):
        embed1.add_field(name=f"**{coordinates[i][0]}**", inline=True, value=set_value(coordinates[i][1], hour))
    embed1.set_footer(text="Made by Marsisus, Magos, Hugo, Gabriel, Antonin...       ")
    embeds.append(embed1)

    if len(coordinates) > 25:
        embed2 = discord.Embed(
                               timestamp=datetime.now())

        for i in range(25, len(coordinates)):
            embed2.add_field(name=f"**{coordinates[i][0]}**", value=set_value(coordinates[i][1], hour), inline=True)
        embed2.set_footer(text="Made by Marsisus, Magos, Hugo, Gabriel, Antonin...       ")
        embeds.append(embed2)

    return embeds